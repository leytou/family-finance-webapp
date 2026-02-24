from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

wb = Workbook()

# Colors
BLUE = "2B6CB0"        # 输入值字体颜色（柔和蓝）
BLACK = "2D3748"       # 公式字体颜色（深灰蓝）
GREEN = "2F855A"       # 链接公式字体颜色（柔和绿）
DARK_BG = "2D3748"     # 深色背景（标题）
MED_BG = "4A90D9"      # 中等背景（章节标题）
LIGHT_BG = "EDF2F7"    # 浅色背景（参数名）
WHITE = "FFFFFF"       # 白色字体
YELLOW_BG = "EBF4FF"   # 淡蓝背景（输入单元格）
GREEN_BG = "E6F4EA"    # 淡绿背景（年终奖月/收入小计）
RED_BG = "FDE8E8"      # 淡红背景（支出小计）
AMBER_BG = "FEF3C7"    # 淡琥珀背景（重大支出月/年度结余）
GRAY_BG = "F7FAFC"     # 极浅灰蓝背景（偶数行/备注）
ORANGE_BG = "FDE8E8"   # 淡红背景（与 RED_BG 统一）

YUAN = '¥#,##0;(¥#,##0);"-"'
MONTH_NAMES = ['','1月','2月','3月','4月','5月','6月','7月','8月','9月','10月','11月','12月']

def hdr(cell, bg=DARK_BG, fg=WHITE, bold=True, sz=11):
    cell.font = Font(bold=bold, color=fg, size=sz)
    cell.fill = PatternFill("solid", start_color=bg)
    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

def inp(cell, bold=False):
    cell.font = Font(color=BLUE, bold=bold)

def fml(cell):
    cell.font = Font(color=BLACK)

def lnk(cell):
    cell.font = Font(color=GREEN)

def bdr(cell):
    s = Side(style='thin')
    cell.border = Border(left=s, right=s, top=s, bottom=s)

def center(cell):
    cell.alignment = Alignment(horizontal='center', vertical='center')

def right(cell):
    cell.alignment = Alignment(horizontal='right', vertical='center')

months_data = []
for y in range(2026, 2031):
    for m in range(2 if y == 2026 else 1, 13):
        months_data.append((y, m))
months_data.append((2031, 1))

def time_lookup_expr(assumption_row, year_cell, mo_num_expr):
    """
    生成时间查找表达式（不含前导 '='），可嵌入更大公式。
    assumption_row: 总览表假设数据行号（6-14）
    year_cell: 当前行年份单元格引用，如 'B3'
    mo_num_expr: 当前行月份数字表达式，如 'VALUE(LEFT(C3,LEN(C3)-1))'
    日期列存储格式为 YYYYMM 整数（如 202702 = 2027年2月），与 year*100+month 比较。
    变更列对：D/E=变更1, F/G=变更2, H/I=变更3, J/K=变更4, L/M=变更5（最后一次优先）
    """
    r = assumption_row
    yyyymm = f'{year_cell}*100+{mo_num_expr}'
    change_cols = [('D','E'), ('F','G'), ('H','I'), ('J','K'), ('L','M')]
    formula = f'总览!$B${r}'
    for date_col, val_col in change_cols:
        date_ref = f'总览!${date_col}${r}'
        val_ref  = f'总览!${val_col}${r}'
        formula = f'IF(AND({date_ref}<>"",{yyyymm}>={date_ref}),{val_ref},{formula})'
    return formula

def time_lookup(assumption_row, year_cell, mo_num_expr):
    """返回完整 Excel 公式（含前导 '='）。"""
    return '=' + time_lookup_expr(assumption_row, year_cell, mo_num_expr)

def get_row_data(year, month):
    has_special = (year == 2026 and month == 5) or (year == 2027 and month == 9) or (month == 10)
    notes = ""
    if year == 2026 and month == 5: notes = "结婚支出"
    elif year == 2027 and month == 9: notes = "孕产费用"
    elif month == 10: notes = "年度旅游"
    return has_special, notes

def create_monthly_sheet(wb, sheet_name, is_first_sheet=True, housing_cfg=None):
    """
    生成月度现金流工作表。
    wb: Workbook 对象
    sheet_name: 工作表名称
    is_first_sheet: 是否为第一个工作表（使用 wb.active）
    housing_cfg: 买房配置 dict，None 表示不买房方案
        {'buy_time_row': 18, 'price_row': 19, 'down_pct_row': 20,
         'loan_years_row': 21, 'loan_rate_row': 22,
         'down_amount_row': 23, 'loan_amount_row': 24, 'monthly_payment_row': 25}
    """
    cl = lambda c: get_column_letter(c)

    # 列号映射
    if housing_cfg:
        # 买房表：18列
        COL_IDX, COL_YEAR, COL_MONTH = 1, 2, 3
        COL_INCOME, COL_BONUS, COL_INVEST, COL_TOTAL_IN = 4, 5, 6, 7
        COL_RENT, COL_DAILY, COL_CHILD = 8, 9, 10
        COL_SPECIAL, COL_NOTE = 11, 12
        COL_DOWN, COL_MORTGAGE = 13, 14
        COL_TOTAL_OUT, COL_NET, COL_CUM = 15, 16, 17
        COL_LOAN_BAL = 18
        all_hdrs = ['序号','年份','月份','月收入(元)','年终奖(元)','理财收益(元)','总收入(元)',
                    '房租(元)','日常开销(元)','孩子开销(元)',
                    '特殊支出(元)','特殊说明',
                    '首付支出(元)','房贷月供(元)',
                    '总支出(元)','当月净储蓄(元)','累计储蓄(元)','房贷剩余(元)']
        all_col_w = [5,7,6,13,13,13,13,10,13,13,13,14,13,13,13,15,15,15]
    else:
        # 不买房表：15列
        COL_IDX, COL_YEAR, COL_MONTH = 1, 2, 3
        COL_INCOME, COL_BONUS, COL_INVEST, COL_TOTAL_IN = 4, 5, 6, 7
        COL_RENT, COL_DAILY, COL_CHILD = 8, 9, 10
        COL_SPECIAL, COL_NOTE = 11, 12
        COL_DOWN, COL_MORTGAGE = None, None
        COL_TOTAL_OUT, COL_NET, COL_CUM = 13, 14, 15
        COL_LOAN_BAL = None
        all_hdrs = ['序号','年份','月份','月收入(元)','年终奖(元)','理财收益(元)','总收入(元)',
                    '房租(元)','日常开销(元)','孩子开销(元)',
                    '特殊支出(元)','特殊说明','总支出(元)','当月净储蓄(元)','累计储蓄(元)']
        all_col_w = [5,7,6,13,13,13,13,10,13,13,13,14,13,15,15]

    if is_first_sheet:
        ws = wb.active
        ws.title = sheet_name
    else:
        ws = wb.create_sheet(sheet_name)

    for i,(h,w) in enumerate(zip(all_hdrs,all_col_w),1):
        c = ws.cell(row=1,column=i,value=h)
        hdr(c)
        ws.column_dimensions[cl(i)].width = w
    ws.row_dimensions[1].height = 42
    ws.freeze_panes = 'A2'

    for idx,(year,month) in enumerate(months_data,1):
        r = idx + 1
        has_special, notes = get_row_data(year, month)

        bg = None
        if has_special: bg = AMBER_BG
        elif month == 12: bg = GREEN_BG
        elif idx % 2 == 0: bg = GRAY_BG

        def sc(col,val,style_fn=None,fmt=None,_bg=bg):
            c = ws.cell(row=r,column=col,value=val)
            if style_fn: style_fn(c)
            if fmt: c.number_format = fmt
            if _bg: c.fill = PatternFill("solid",start_color=_bg)
            c.alignment = Alignment(horizontal='center' if col<=3 else 'right',vertical='center')
            return c

        # 月份判断辅助：从月份名提取数字，如"9月"→9
        mo_num = f'VALUE(LEFT({cl(COL_MONTH)}{r},LEN({cl(COL_MONTH)}{r})-1))'
        after_child_cond = f'OR({cl(COL_YEAR)}{r}>2027,AND({cl(COL_YEAR)}{r}=2027,{mo_num}>=9))'

        sc(COL_IDX,idx); sc(COL_YEAR,year); sc(COL_MONTH,MONTH_NAMES[month])

        # 月收入：时间查找（总览 row 6）
        sc(COL_INCOME, time_lookup(6, f'{cl(COL_YEAR)}{r}', mo_num), fml, YUAN)

        # 年终奖：12月才有，金额时间查找（总览 row 7）
        bonus_lookup = time_lookup_expr(7, f'{cl(COL_YEAR)}{r}', mo_num)
        sc(COL_BONUS, f'=IF({cl(COL_MONTH)}{r}="12月",{bonus_lookup},0)', fml, YUAN)

        # 理财收益：上月累计储蓄 × 年利率（总览 row 14）/ 12
        if idx == 1:
            sc(COL_INVEST, f'=ROUND(总览!$B$8*总览!$B$14/12,0)', fml, YUAN)
        else:
            sc(COL_INVEST, f'=ROUND({cl(COL_CUM)}{r-1}*总览!$B$14/12,0)', fml, YUAN)

        # 总收入
        sc(COL_TOTAL_IN, f'={cl(COL_INCOME)}{r}+{cl(COL_BONUS)}{r}+{cl(COL_INVEST)}{r}', fml, YUAN)

        # 房租：生育前用婚前租金(row9)，生育后用生育后租金(row10)
        rent_before = time_lookup_expr(9,  f'{cl(COL_YEAR)}{r}', mo_num)
        rent_after  = time_lookup_expr(10, f'{cl(COL_YEAR)}{r}', mo_num)
        sc(COL_RENT, f'=IF({after_child_cond},{rent_after},{rent_before})', fml, YUAN)

        # 日常开销：时间查找（总览 row 11）
        sc(COL_DAILY, time_lookup(11, f'{cl(COL_YEAR)}{r}', mo_num), fml, YUAN)

        # 孩子月开销：生育后才有，时间查找（总览 row 12）
        child_val = time_lookup_expr(12, f'{cl(COL_YEAR)}{r}', mo_num)
        sc(COL_CHILD, f'=IF({after_child_cond},{child_val},0)', fml, YUAN)

        # 特殊支出：年度旅游引用总览 row 13（时间查找）
        travel_val = time_lookup_expr(13, f'{cl(COL_YEAR)}{r}', mo_num)
        sc(COL_SPECIAL, f'=IF(AND({cl(COL_YEAR)}{r}=2026,{cl(COL_MONTH)}{r}="5月"),120000,'
               f'IF(AND({cl(COL_YEAR)}{r}=2027,{cl(COL_MONTH)}{r}="9月"),30000,'
               f'IF({cl(COL_MONTH)}{r}="10月",{travel_val},0)))', fml, YUAN)

        # 特殊说明
        nc = ws.cell(row=r,column=COL_NOTE,value=notes)
        nc.alignment = Alignment(horizontal='center',vertical='center')
        if bg: nc.fill = PatternFill("solid",start_color=bg)

        # 买房专属列
        if housing_cfg:
            hc = housing_cfg
            yyyymm = f'{cl(COL_YEAR)}{r}*100+{mo_num}'
            buy_time = f'总览!$B${hc["buy_time_row"]}'
            down_amt = f'总览!$B${hc["down_amount_row"]}'
            loan_amt = f'总览!$B${hc["loan_amount_row"]}'
            monthly_pmt = f'总览!$B${hc["monthly_payment_row"]}'
            loan_rate = f'总览!$B${hc["loan_rate_row"]}'

            # 首付支出：仅买房当月
            sc(COL_DOWN, f'=IF({yyyymm}={buy_time},{down_amt},0)', fml, YUAN)

            # 房贷月供：买房次月起
            sc(COL_MORTGAGE, f'=IF({yyyymm}>{buy_time},{monthly_pmt},0)', fml, YUAN)

            # 房贷剩余
            if idx == 1:
                sc(COL_LOAN_BAL,
                   f'=IF({yyyymm}<{buy_time},0,'
                   f'IF({yyyymm}={buy_time},{loan_amt},{loan_amt}))',
                   fml, YUAN)
            else:
                prev_bal = f'{cl(COL_LOAN_BAL)}{r-1}'
                sc(COL_LOAN_BAL,
                   f'=IF({yyyymm}<{buy_time},0,'
                   f'IF({yyyymm}={buy_time},{loan_amt},'
                   f'{prev_bal}-({monthly_pmt}-{prev_bal}*{loan_rate}/12)))',
                   fml, YUAN)

            # 总支出（含首付和月供）
            sc(COL_TOTAL_OUT,
               f'=SUM({cl(COL_RENT)}{r}:{cl(COL_CHILD)}{r})+{cl(COL_SPECIAL)}{r}+{cl(COL_DOWN)}{r}+{cl(COL_MORTGAGE)}{r}',
               fml, YUAN)
        else:
            # 不买房：总支出 = SUM(房租:特殊支出)
            sc(COL_TOTAL_OUT, f'=SUM({cl(COL_RENT)}{r}:{cl(COL_SPECIAL)}{r})', fml, YUAN)

        # 当月净储蓄
        sc(COL_NET, f'={cl(COL_TOTAL_IN)}{r}-{cl(COL_TOTAL_OUT)}{r}', fml, YUAN)

        # 累计储蓄
        if idx==1:
            sc(COL_CUM, f'=总览!$B$8+{cl(COL_NET)}{r}', fml, YUAN)
        else:
            sc(COL_CUM, f'={cl(COL_CUM)}{r-1}+{cl(COL_NET)}{r}', fml, YUAN)

    # 图例行
    lr = len(months_data)+3
    ws.cell(row=lr,column=1,value='图例说明：').font = Font(bold=True)
    for col,txt,bg_c in [(2,'蓝色=输入值',LIGHT_BG),(4,'黄底=重大支出月',AMBER_BG),(6,'绿底=年终奖月',GREEN_BG)]:
        c = ws.cell(row=lr,column=col,value=txt)
        c.fill = PatternFill("solid",start_color=bg_c)
        c.font = Font(color=BLUE if bg_c==LIGHT_BG else BLACK)

    return ws

# ============================================================
# Sheet 2: 月度现金流（不买房）
# ============================================================
ws2 = create_monthly_sheet(wb, "月度现金流", is_first_sheet=True, housing_cfg=None)

# ============================================================
# Sheet 3: 月度现金流（买房）
# ============================================================
HOUSING_CFG = {
    'buy_time_row': 18,
    'price_row': 19,
    'down_pct_row': 20,
    'loan_years_row': 21,
    'loan_rate_row': 22,
    'down_amount_row': 23,
    'loan_amount_row': 24,
    'monthly_payment_row': 25,
}
ws3 = create_monthly_sheet(wb, "月度现金流（买房）", is_first_sheet=False, housing_cfg=HOUSING_CFG)

# ============================================================
# Sheet 1: 总览 (insert before monthly sheet)
# ============================================================
ws1 = wb.create_sheet("总览", 0)

# Title
ws1.merge_cells('A1:L1')
ws1['A1'] = '家庭5年财务规划（2026-2030）'
ws1['A1'].font = Font(bold=True, size=16, color=WHITE)
ws1['A1'].fill = PatternFill("solid", start_color=DARK_BG)
ws1['A1'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[1].height = 40

ws1.merge_cells('A2:L2')
ws1['A2'] = '制定日期：2026年2月  |  规划周期：2026年2月 - 2030年12月'
ws1['A2'].font = Font(size=10, color="718096")
ws1['A2'].alignment = Alignment(horizontal='center', vertical='center')
ws1.row_dimensions[2].height = 22

# Section: 基本假设
ws1.merge_cells('A4:M4')
ws1['A4'] = '📋 基本财务假设'
hdr(ws1['A4'], bg=MED_BG, sz=12)
ws1.row_dimensions[4].height = 28

# 子标题行 row 5
sub_hdrs = [
    '参数名', '初始值', '备注',
    '变更1生效\n(YYYYMM)', '变更1值',
    '变更2生效\n(YYYYMM)', '变更2值',
    '变更3生效\n(YYYYMM)', '变更3值',
    '变更4生效\n(YYYYMM)', '变更4值',
    '变更5生效\n(YYYYMM)', '变更5值',
]
sub_col_w = [22, 16, 22, 14, 14, 14, 14, 14, 14, 14, 14, 14, 14]
assert len(sub_hdrs) == len(sub_col_w), "列标题与列宽数量不匹配"
for i, (h, w) in enumerate(zip(sub_hdrs, sub_col_w), 1):
    c = ws1.cell(row=5, column=i, value=h)
    hdr(c, bg='4A90D9', sz=9)
    ws1.column_dimensions[get_column_letter(i)].width = w
ws1.row_dimensions[5].height = 30

# 假设数据从 row 6 开始（row 5 为子标题行）
assumptions = [
    ('双方月可支配收入合计', 26000, YUAN, '每月税后可支配收入'),
    ('年终奖（合计）',       46000, YUAN, ''),
    ('当前存款',             380000, YUAN, '截至2026年2月'),
    ('婚前月租金',           2400,  YUAN, '生育前，含停车费'),
    ('生育后月租金',         3900,  YUAN, '孩子出生后搬入更大住所，含停车费'),
    ('月日常开销',           1500,  YUAN, '餐饮/交通/娱乐等'),
    ('孩子月开销',           3000,  YUAN, '奶粉/纸尿裤/早教等（出生后）'),
    ('年度旅游支出',         15000, YUAN, ''),
    ('存款年利率',           0.025, '0.0%', '按年初余额计算，保守型理财参考收益率'),
]

for i, (label, val, fmt, note) in enumerate(assumptions, 6):
    ws1.row_dimensions[i].height = 22
    c_label = ws1.cell(row=i, column=1, value=label)
    c_label.font = Font(bold=True)
    c_label.fill = PatternFill('solid', start_color=LIGHT_BG)
    c_label.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    c_val = ws1.cell(row=i, column=2, value=val)
    inp(c_val)
    c_val.number_format = fmt
    c_val.alignment = Alignment(horizontal='right', vertical='center')
    c_val.fill = PatternFill('solid', start_color=YELLOW_BG)

    c_note = ws1.cell(row=i, column=3, value=note)
    c_note.font = Font(color='718096', size=9)
    c_note.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    # 变更列 D-M：黄底蓝字输入格（空值）
    for col in range(4, 14):
        c = ws1.cell(row=i, column=col, value=None)
        c.fill = PatternFill('solid', start_color=YELLOW_BG)
        inp(c)
        c.alignment = Alignment(horizontal='right', vertical='center')
        # 偶数列号 = D,F,H,J,L = 年月列，奇数列号 = E,G,I,K,M = 值列
        if col % 2 == 0:  # 偶数列号 = 年月列
            c.number_format = '0'
        else:             # 奇数列号 = 值列
            c.number_format = fmt

# ============================================================
# Section: 买房假设
# ============================================================
ws1.merge_cells('A16:M16')
ws1['A16'] = '🏠 买房假设'
hdr(ws1['A16'], bg=MED_BG, sz=12)
ws1.row_dimensions[16].height = 28

# 买房假设子标题行 row 17
for i, (h, w) in enumerate(zip(sub_hdrs, sub_col_w), 1):
    c = ws1.cell(row=17, column=i, value=h)
    hdr(c, bg='4A90D9', sz=9)
ws1.row_dimensions[17].height = 30

# 买房假设参数（row 18-22）
housing_assumptions = [
    ('计划买房时间',   202801, '0',   'YYYYMM格式，如202801=2028年1月'),
    ('总房价',         1500000, YUAN,  ''),
    ('首付比例',       0.3,    '0%',  ''),
    ('贷款年限',       30,     '0',   '年'),
    ('贷款年利率',     0.031,  '0.0%','商贷参考利率'),
]

for i, (label, val, fmt, note) in enumerate(housing_assumptions, 18):
    ws1.row_dimensions[i].height = 22
    c_label = ws1.cell(row=i, column=1, value=label)
    c_label.font = Font(bold=True)
    c_label.fill = PatternFill('solid', start_color=LIGHT_BG)
    c_label.alignment = Alignment(horizontal='left', vertical='center', indent=1)

    c_val = ws1.cell(row=i, column=2, value=val)
    inp(c_val)
    c_val.number_format = fmt
    c_val.alignment = Alignment(horizontal='right', vertical='center')
    c_val.fill = PatternFill('solid', start_color=YELLOW_BG)

    c_note = ws1.cell(row=i, column=3, value=note)
    c_note.font = Font(color='718096', size=9)
    c_note.alignment = Alignment(horizontal='left', vertical='center', indent=1)

# 计算行（row 23-25）
calc_rows = [
    (23, '首付金额', '=B19*B20', '自动计算'),
    (24, '贷款总额', '=B19-B23', '自动计算'),
    (25, '月供金额', '=-PMT(B22/12,B21*12,B24)', '等额本息，自动计算'),
]
for row_num, label, formula, note in calc_rows:
    ws1.row_dimensions[row_num].height = 22
    c_label = ws1.cell(row=row_num, column=1, value=label)
    c_label.font = Font(bold=True)
    c_label.fill = PatternFill('solid', start_color=LIGHT_BG)
    c_label.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c_val = ws1.cell(row=row_num, column=2, value=formula)
    fml(c_val)
    c_val.number_format = YUAN
    c_val.alignment = Alignment(horizontal='right', vertical='center')
    c_note = ws1.cell(row=row_num, column=3, value=note)
    c_note.font = Font(color='718096', size=9)
    c_note.alignment = Alignment(horizontal='left', vertical='center', indent=1)

# ============================================================
# Section: 五年收支总览（并排对比）
# ============================================================
SUMMARY_START = 27
R_HDR = SUMMARY_START + 1
R_YEAR_HDR = R_HDR + 1
RD = R_YEAR_HDR + 1  # row 30，数据起始行

ws1.merge_cells(f'A{SUMMARY_START}:L{SUMMARY_START}')
ws1.cell(row=SUMMARY_START, column=1, value='📊 五年收支总览（2026—2030）')
hdr(ws1.cell(row=SUMMARY_START, column=1), bg=MED_BG, sz=12)
ws1.row_dimensions[SUMMARY_START].height = 28

# 方案标注行
ws1.merge_cells(f'B{R_HDR}:F{R_HDR}')
c = ws1.cell(row=R_HDR, column=2, value='📋 不买房方案')
hdr(c, bg="4A90D9", sz=10)
ws1.column_dimensions['G'].width = 3
ws1.merge_cells(f'H{R_HDR}:L{R_HDR}')
c = ws1.cell(row=R_HDR, column=8, value='🏠 买房方案')
hdr(c, bg="4A90D9", sz=10)
c = ws1.cell(row=R_HDR, column=1, value='')
hdr(c, bg="4A90D9", sz=10)
ws1.row_dimensions[R_HDR].height = 24

# 年份列标题行
years_hdrs = ['项目','2026年','2027年','2028年','2029年','2030年']
for i,h in enumerate(years_hdrs, 1):
    c = ws1.cell(row=R_YEAR_HDR, column=i, value=h)
    hdr(c, bg="4A90D9", sz=10)
for i,h in enumerate(['2026年','2027年','2028年','2029年','2030年'], 8):
    c = ws1.cell(row=R_YEAR_HDR, column=i, value=h)
    hdr(c, bg="4A90D9", sz=10)
ws1.row_dimensions[R_YEAR_HDR].height = 24

# 行号变量
ROW_BAL_START = RD + 0
ROW_INCOME_SEC = RD + 1
ROW_SALARY = RD + 2
ROW_BONUS_S = RD + 3
ROW_INVEST_S = RD + 4
ROW_INCOME_TOTAL = RD + 5
ROW_EXPENSE_SEC = RD + 6
ROW_RENT_S = RD + 7
ROW_DAILY_S = RD + 8
ROW_CHILD_S = RD + 9
ROW_TRAVEL_S = RD + 10
ROW_DOWN_PAY = RD + 11
ROW_MORTGAGE_PAY = RD + 12
ROW_SPECIAL_SEC = RD + 13
ROW_WEDDING = RD + 14
ROW_BABY = RD + 15
ROW_NOTE_ROW = RD + 16
ROW_EXP_TOTAL = RD + 17
ROW_EMPTY = RD + 18
ROW_ANNUAL_NET = RD + 19
ROW_BAL_END = RD + 20

ROW_DEFS = [
    (ROW_BAL_START,'年初余额',True,LIGHT_BG,False),
    (ROW_INCOME_SEC,'【收入】',True,None,True),
    (ROW_SALARY,'  工资收入',False,None,False),
    (ROW_BONUS_S,'  年终奖',False,None,False),
    (ROW_INVEST_S,'  理财收益',False,None,False),
    (ROW_INCOME_TOTAL,'收入小计',True,GREEN_BG,False),
    (ROW_EXPENSE_SEC,'【支出】',True,None,True),
    (ROW_RENT_S,'  房租',False,None,False),
    (ROW_DAILY_S,'  日常开销',False,None,False),
    (ROW_CHILD_S,'  育儿开销',False,None,False),
    (ROW_TRAVEL_S,'  旅游支出',False,None,False),
    (ROW_DOWN_PAY,'  首付支出',False,None,False),
    (ROW_MORTGAGE_PAY,'  房贷月供',False,None,False),
    (ROW_SPECIAL_SEC,'  【特殊支出】',True,'E2E8F0',False),
    (ROW_WEDDING,'    结婚',False,None,False),
    (ROW_BABY,'    产检及分娩',False,None,False),
    (ROW_NOTE_ROW,'',False,GRAY_BG,False),
    (ROW_EXP_TOTAL,'支出小计',True,RED_BG,False),
    (ROW_EMPTY,'',False,None,False),
    (ROW_ANNUAL_NET,'年度结余（收入−支出）',True,AMBER_BG,False),
    (ROW_BAL_END,'年末累计余额',True,GREEN_BG,False),
]

SECT_HDR_BG = "E2E8F0"
for row,label,bold,bg_c,is_sect in ROW_DEFS:
    ws1.row_dimensions[row].height = 22
    c = ws1.cell(row=row,column=1,value=label)
    if is_sect:
        ws1.merge_cells(f'A{row}:L{row}')
        c.font = Font(bold=True,size=10,color="2D3748")
        c.fill = PatternFill("solid",start_color=SECT_HDR_BG)
        c.alignment = Alignment(horizontal='left',vertical='center',indent=1)
    else:
        c.font = Font(bold=bold,size=10)
        if bg_c: c.fill = PatternFill("solid",start_color=bg_c)
        c.alignment = Alignment(horizontal='left',vertical='center',indent=1)

def fill_cell(ws, row, col, val, style_fn=None, fmt=YUAN, bold=False, bg_c=None):
    c = ws.cell(row=row, column=col, value=val)
    if style_fn: style_fn(c)
    if bold: c.font = Font(bold=True, color=c.font.color if c.font.color else BLACK)
    if fmt: c.number_format = fmt
    if bg_c: c.fill = PatternFill("solid", start_color=bg_c)
    c.alignment = Alignment(horizontal='right', vertical='center')
    return c

# 不买房方案（B-F列，引用"月度现金流"表）
SH_NO = '月度现金流'
for yi,year in enumerate([2026,2027,2028,2029,2030]):
    col = yi + 2
    cl = get_column_letter(col)

    if year == 2026:
        fill_cell(ws1, ROW_BAL_START, col, '=$B$8', fml, bg_c=YELLOW_BG)
    else:
        pcl = get_column_letter(col-1)
        fill_cell(ws1, ROW_BAL_START, col, f'={pcl}{ROW_BAL_END}', fml)

    fill_cell(ws1, ROW_SALARY, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!D:D)', lnk)
    fill_cell(ws1, ROW_BONUS_S, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!E:E)', lnk)
    fill_cell(ws1, ROW_INVEST_S, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!F:F)', lnk)
    fill_cell(ws1, ROW_INCOME_TOTAL, col, f'=SUM({cl}{ROW_SALARY}:{cl}{ROW_INVEST_S})', fml, bold=True, bg_c=GREEN_BG)

    fill_cell(ws1, ROW_RENT_S, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!H:H)', lnk)
    fill_cell(ws1, ROW_DAILY_S, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!I:I)', lnk)
    fill_cell(ws1, ROW_CHILD_S, col, f'=SUMIF({SH_NO}!B:B,{year},{SH_NO}!J:J)', lnk)
    fill_cell(ws1, ROW_TRAVEL_S, col, f'=SUMIFS({SH_NO}!K:K,{SH_NO}!B:B,{year},{SH_NO}!L:L,"年度旅游")', lnk)
    fill_cell(ws1, ROW_WEDDING, col, f'=SUMIFS({SH_NO}!K:K,{SH_NO}!B:B,{year},{SH_NO}!L:L,"结婚支出")', lnk)
    fill_cell(ws1, ROW_BABY, col, f'=SUMIFS({SH_NO}!K:K,{SH_NO}!B:B,{year},{SH_NO}!L:L,"孕产费用")', lnk)
    fill_cell(ws1, ROW_EXP_TOTAL, col,
              f'=SUM({cl}{ROW_RENT_S}:{cl}{ROW_TRAVEL_S})+SUM({cl}{ROW_WEDDING}:{cl}{ROW_BABY})',
              fml, bold=True, bg_c=RED_BG)

    fill_cell(ws1, ROW_ANNUAL_NET, col, f'={cl}{ROW_INCOME_TOTAL}-{cl}{ROW_EXP_TOTAL}', fml, bold=True, bg_c=AMBER_BG)
    fill_cell(ws1, ROW_BAL_END, col, f'={cl}{ROW_BAL_START}+{cl}{ROW_ANNUAL_NET}', fml, bold=True, bg_c=GREEN_BG)

# 买房方案（H-L列，引用"月度现金流（买房）"表）
SH_BUY = '月度现金流（买房）'
for yi,year in enumerate([2026,2027,2028,2029,2030]):
    col = yi + 8
    cl = get_column_letter(col)

    if year == 2026:
        fill_cell(ws1, ROW_BAL_START, col, '=$B$8', fml, bg_c=YELLOW_BG)
    else:
        pcl = get_column_letter(col-1)
        fill_cell(ws1, ROW_BAL_START, col, f'={pcl}{ROW_BAL_END}', fml)

    fill_cell(ws1, ROW_SALARY, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!D:D)', lnk)
    fill_cell(ws1, ROW_BONUS_S, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!E:E)', lnk)
    fill_cell(ws1, ROW_INVEST_S, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!F:F)', lnk)
    fill_cell(ws1, ROW_INCOME_TOTAL, col, f'=SUM({cl}{ROW_SALARY}:{cl}{ROW_INVEST_S})', fml, bold=True, bg_c=GREEN_BG)

    fill_cell(ws1, ROW_RENT_S, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!H:H)', lnk)
    fill_cell(ws1, ROW_DAILY_S, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!I:I)', lnk)
    fill_cell(ws1, ROW_CHILD_S, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!J:J)', lnk)
    fill_cell(ws1, ROW_TRAVEL_S, col, f'=SUMIFS(\'{SH_BUY}\'!K:K,\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!L:L,"年度旅游")', lnk)
    fill_cell(ws1, ROW_DOWN_PAY, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!M:M)', lnk)
    fill_cell(ws1, ROW_MORTGAGE_PAY, col, f'=SUMIF(\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!N:N)', lnk)
    fill_cell(ws1, ROW_WEDDING, col, f'=SUMIFS(\'{SH_BUY}\'!K:K,\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!L:L,"结婚支出")', lnk)
    fill_cell(ws1, ROW_BABY, col, f'=SUMIFS(\'{SH_BUY}\'!K:K,\'{SH_BUY}\'!B:B,{year},\'{SH_BUY}\'!L:L,"孕产费用")', lnk)
    fill_cell(ws1, ROW_EXP_TOTAL, col,
              f'=SUM({cl}{ROW_RENT_S}:{cl}{ROW_TRAVEL_S})+{cl}{ROW_DOWN_PAY}+{cl}{ROW_MORTGAGE_PAY}+SUM({cl}{ROW_WEDDING}:{cl}{ROW_BABY})',
              fml, bold=True, bg_c=RED_BG)

    fill_cell(ws1, ROW_ANNUAL_NET, col, f'={cl}{ROW_INCOME_TOTAL}-{cl}{ROW_EXP_TOTAL}', fml, bold=True, bg_c=AMBER_BG)
    fill_cell(ws1, ROW_BAL_END, col, f'={cl}{ROW_BAL_START}+{cl}{ROW_ANNUAL_NET}', fml, bold=True, bg_c=GREEN_BG)

# 备注行
ws1.merge_cells(f'A{ROW_NOTE_ROW}:L{ROW_NOTE_ROW}')
note_cell = ws1.cell(row=ROW_NOTE_ROW, column=1)
note_cell.value = '* 特殊支出数据来源：月度现金流汇总；结婚（2026年5月）及孕产费用（2027年9月）均为一次性支出'
note_cell.font = Font(size=9, color="718096", italic=True)
note_cell.fill = PatternFill("solid", start_color=GRAY_BG)
note_cell.alignment = Alignment(horizontal='left', vertical='center', indent=1, wrap_text=True)
ws1.row_dimensions[ROW_NOTE_ROW].height = 28

# 添加边框（不买房 A-F + 买房 H-L）
thin = Side(style='thin')
for row in range(R_YEAR_HDR, ROW_BAL_END+1):
    for col in list(range(1,7)) + list(range(8,13)):
        ws1.cell(row=row,column=col).border = Border(left=thin,right=thin,top=thin,bottom=thin)

# ============================================================
# Save
# ============================================================
output_path = 'C:/Users/10204/Desktop/plan/五年规划.xlsx'
wb.save(output_path)
print(f"Saved: {output_path}")
